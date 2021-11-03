import time
import os, json
from termcolor import colored
from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# GLOBAL VARIABLES
BASE_FOLDER = os.getcwd()
PATH_TO_CHROME_DRIVER = os.path.join(BASE_FOLDER, 'chromedriver')
PATH_TO_EXCEL_DATA_FILE = os.path.join(BASE_FOLDER, 'data.xlsx')

# FUNCTIONS
def initalizeDriver():
    options = Options()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_extension('metamask.crx')
    driver = webdriver.Chrome(executable_path=PATH_TO_CHROME_DRIVER, options=options)
    wait = WebDriverWait(driver, 60)
    return driver, wait

def fileExists(path=PATH_TO_EXCEL_DATA_FILE):
    return os.path.exists(PATH_TO_EXCEL_DATA_FILE)

def readData(path=PATH_TO_EXCEL_DATA_FILE):
    wb = load_workbook(PATH_TO_EXCEL_DATA_FILE)
    ws = wb.active
    data = list()
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = row[0]
        metadata = row[1]
        d = {'url': url, 'metadata': metadata}
        data.append(d)
    return data

def waitForLogin():
    txt = "Please login to Cargo website."
    color = "green"
    text = colored(txt, color)
    print(text)
    condition = True
    while (condition):
        val = input("Please 'Y/y' key again to continue: ")
        if val == 'y' or val == 'Y':
            condition = False

def scarpe(driver, wait, url, metadata):
    driver.get(url)
    # Side Bar
    wait.until(EC.presence_of_element_located((By.CLASS_NAME, "MuiPaper-root")))
    # Sell / Edit
    wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(@class, 'MuiBox-root') and p[contains(., 'Sell/Edit')]]")))
    btn_xpath = driver.find_element_by_xpath("//div[contains(@class, 'MuiBox-root') and p[contains(., 'Sell/Edit')]]")
    btn_xpath.click()
    # Text area
    wait.until(EC.presence_of_element_located((By.NAME, "editor")))
    editor = driver.find_element_by_name("editor")
    editor.clear()
    editor.send_keys(metadata)
    # Save meta changes button
    btn_xpath = "//button[@data-id='cargo-btn' and contains(text(), 'Save metadata changes')]"
    wait.until(EC.presence_of_element_located((By.XPATH, btn_xpath)))
    btn = driver.find_element_by_xpath(btn_xpath)
    btn.click()
    time.sleep(2)